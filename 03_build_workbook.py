import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference

BASE = "/home/claude/FUTURE_DS_01/data"
cleaned = pd.read_csv(f"{BASE}/cleaned_sales_data.csv")
monthly = pd.read_csv(f"{BASE}/monthly_trend.csv")
top_products = pd.read_csv(f"{BASE}/top_products.csv")
cat_perf = pd.read_csv(f"{BASE}/category_performance.csv")
region_perf = pd.read_csv(f"{BASE}/region_performance.csv")

HEADER_FILL = PatternFill("solid", fgColor="2E5EAA")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="2E5EAA")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_df(ws, df, start_row=1, start_col=1, title=None):
    r = start_row
    if title:
        ws.cell(row=r, column=start_col, value=title).font = TITLE_FONT
        r += 2
    for j, col in enumerate(df.columns):
        c = ws.cell(row=r, column=start_col + j, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
    for i, row in enumerate(df.itertuples(index=False), start=1):
        for j, val in enumerate(row):
            c = ws.cell(row=r + i, column=start_col + j, value=val)
            c.font = BODY_FONT
            c.border = BORDER
    for j, col in enumerate(df.columns):
        col_letter = get_column_letter(start_col + j)
        maxlen = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str)])
        ws.column_dimensions[col_letter].width = min(max(maxlen + 3, 10), 40)
    return r + len(df) + 1

wb = Workbook()

# --- Sheet 1: Raw Data (cleaned) ---
ws1 = wb.active
ws1.title = "Cleaned Data"
write_df(ws1, cleaned)
ws1.freeze_panes = "A2"

# --- Sheet 2: KPI Summary (formula-driven) ---
ws2 = wb.create_sheet("KPI Summary")
ws2.cell(row=1, column=1, value="Business Sales Performance - KPI Summary").font = TITLE_FONT
n = len(cleaned)
last_row = n + 1  # data sheet has header row 1, data rows 2..n+1
labels = ["Total Revenue", "Total Profit", "Overall Profit Margin %", "Total Orders (unique)", "Average Order Value"]
ws2.cell(row=3, column=1, value="Metric").font = HEADER_FONT
ws2.cell(row=3, column=1).fill = HEADER_FILL
ws2.cell(row=3, column=2, value="Value").font = HEADER_FONT
ws2.cell(row=3, column=2).fill = HEADER_FILL

# find column letters in Cleaned Data
cols = list(cleaned.columns)
def col_letter(name):
    return get_column_letter(cols.index(name) + 1)

sales_col = col_letter("Sales")
profit_col = col_letter("Profit")
orderid_col = col_letter("Order ID")

formulas = [
    f"=SUM('Cleaned Data'!{sales_col}2:{sales_col}{last_row})",
    f"=SUM('Cleaned Data'!{profit_col}2:{profit_col}{last_row})",
    f"=B5/B4*100",
    f"=SUMPRODUCT(1/COUNTIF('Cleaned Data'!{orderid_col}2:{orderid_col}{last_row},'Cleaned Data'!{orderid_col}2:{orderid_col}{last_row}))",
    f"=B4/B7",
]
for i, (lbl, f) in enumerate(zip(labels, formulas)):
    r = 4 + i
    ws2.cell(row=r, column=1, value=lbl).font = BODY_FONT
    c = ws2.cell(row=r, column=2, value=f)
    c.font = Font(name="Arial", size=10, bold=True)
    if "Margin" not in lbl and "Orders" not in lbl:
        c.number_format = '"Rs "#,##0'
    elif "Margin" in lbl:
        c.number_format = '0.00"%"'
    else:
        c.number_format = '#,##0'
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 20

# --- Sheet 3: Category Performance ---
ws3 = wb.create_sheet("Category Performance")
write_df(ws3, cat_perf, title="Revenue, Profit & Margin by Category")

# --- Sheet 4: Region Performance ---
ws4 = wb.create_sheet("Region Performance")
write_df(ws4, region_perf, title="Revenue, Profit & Margin by Region")

# --- Sheet 5: Top Products ---
ws5 = wb.create_sheet("Top Products")
write_df(ws5, top_products.head(20), title="Top 20 Products by Revenue")

# --- Sheet 6: Monthly Trend + native chart ---
ws6 = wb.create_sheet("Monthly Trend")
end_row = write_df(ws6, monthly, title="Monthly Revenue & Profit Trend")
chart = LineChart()
chart.title = "Monthly Revenue Trend"
chart.y_axis.title = "Revenue (Rs)"
chart.x_axis.title = "Month"
data_start = 4
data_end = 3 + len(monthly)
data_ref = Reference(ws6, min_col=2, max_col=3, min_row=3, max_row=data_end)
cats_ref = Reference(ws6, min_col=1, min_row=4, max_row=data_end)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.width = 24
chart.height = 12
ws6.add_chart(chart, "F3")

# --- Sheet 7: Insights & Recommendations ---
ws7 = wb.create_sheet("Insights & Recommendations")
ws7.cell(row=1, column=1, value="Business Sales Performance Analytics - Insights").font = TITLE_FONT
insights = [
    ("Revenue driver", "Furniture and Electronics together generate ~87% of total revenue (Rs 68.7L + Rs 60.1L of Rs 148.7L total). Office Supplies is high in order count (969 orders) but contributes under 3% of revenue - it is a low-value, high-frequency category."),
    ("Best-margin category", "Electronics carries the strongest profit margin (19.99%), narrowly ahead of Furniture (19.71%). Clothing lags at 15.79% margin and should be reviewed for discounting practices."),
    ("Top products", "Office Chair, Wireless Mouse, and Bluetooth Speaker are the top 3 revenue generators. Office Chair also has the highest profit contribution among top products (~Rs 3.94L)."),
    ("Regional performance", "North region leads in both revenue (Rs 44.5L) and margin (19.72%). South, while lowest in revenue, holds a comparable margin (19.36%) - suggesting a demand/marketing gap rather than a profitability problem."),
    ("Seasonality", "Revenue rises sharply in Oct-Dec (festive season, ~1.6x normal demand) and has a smaller July peak (~1.3x, mid-year sale period). Inventory and marketing spend should be planned around these windows."),
    ("Recommendation 1", "Increase marketing/ad spend in the South region during Oct-Dec to close the revenue gap with North, since margins there are already healthy."),
    ("Recommendation 2", "Bundle low-value, high-frequency Office Supplies items with Electronics/Furniture purchases to lift average order value without extra acquisition cost."),
    ("Recommendation 3", "Audit discounting on Clothing - it has the highest average discount usage and the lowest margin of all categories; tightening discount tiers could recover 2-3 margin points."),
    ("Recommendation 4", "Double down on Office Chair and top electronics SKUs with pre-festive season stock build-up and featured placement, since they are simultaneously top-revenue and top-margin."),
]
r = 3
for label, text in insights:
    ws7.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=True, size=10, color="2E5EAA")
    ws7.cell(row=r, column=2, value=text).font = BODY_FONT
    ws7.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws7.row_dimensions[r].height = 45
    r += 1
ws7.column_dimensions["A"].width = 20
ws7.column_dimensions["B"].width = 100

wb.save("/home/claude/FUTURE_DS_01/Task1_Business_Sales_Performance_Analytics.xlsx")
print("Workbook saved.")
